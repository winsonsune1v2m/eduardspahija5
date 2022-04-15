from __future__ import absolute_import, unicode_literals

import json
import os
import redis
from statics.scripts import encryption,read_excel
from mtrops_v2.settings import SECRET_KEY,BASE_DIR,DATABASES
from django.shortcuts import render,HttpResponse,redirect
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from app_asset import models as asset_db
from app_auth import models as auth_db
from app_log import models as log_db
from app_auth.views import login_check,perms_check
from mtrops_v2.settings import SERVER_TAG,SALT_API,WEBSSH_URL,REDIS_INFO
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color,colors,GradientFill,NamedStyle
from app_asset.tasks import sync_host



# Create your views here.



class IDC(View):
    """机房管理"""
    @method_decorator(csrf_exempt)
    @method_decorator(login_check)
    @method_decorator(perms_check)
    def dispatch(self, request, *args, **kwargs):
        return super(IDC,self).dispatch(request, *args, **kwargs)

    def get(self,request):
        title = "IDC 机房"
        idc_obj = asset_db.IDC.objects.all()
        return render(request,'asset_idc.html',locals())

    def post(self,request):
        idc_name = request.POST.get("idc_name")
        idc_msg = request.POST.get("idc_msg")
        idc_admin = request.POST.get("idc_admin")
        idc_admin_phone = request.POST.get("idc_admin_phone")
        idc_admin_email = request.POST.get("idc_admin_email")
        idc_obj = asset_db.IDC(idc_name=idc_name,idc_msg=idc_msg,idc_admin=idc_admin,idc_admin_phone=idc_admin_phone,
                               idc_admin_email=idc_admin_email)
        idc_obj.save()
        data = '机房已添加，请刷新查看！'
        return HttpResponse(data)

    def put(self,request):
        req_info = eval(request.body.decode())
        idc_id = req_info.get("idc_id")
        idc_name = req_info.get("idc_name")
        idc_msg = req_info.get("idc_msg")
        idc_admin = req_info.get("idc_admin")
        idc_admin_phone = req_info.get("idc_admin_phone")
        idc_admin_email = req_info.get("idc_admin_email")
        action = req_info.get("action",None)

        if action:
            """修改IDC信息"""
            idc_obj = asset_db.IDC.objects.get(id=idc_id)
            idc_obj.idc_name = idc_name
            idc_obj.idc_msg = idc_msg
            idc_obj.idc_admin = idc_admin
            idc_obj.idc_admin_phone = idc_admin_phone
            idc_obj.idc_admin_email = idc_admin_email
            idc_obj.save()
            data = "IDC已修改，请刷新查看！"
            return HttpResponse(data)
        else:
            """获取修改信息"""
            idc_info = asset_db.IDC.objects.get(id=idc_id)
            info_json = {'idc_id': idc_info.id, 'idc_name': idc_info.idc_name, 'idc_msg': idc_info.idc_msg,
                         'idc_admin': idc_info.idc_admin,'idc_admin_phone': idc_info.idc_admin_phone,'idc_admin_email': idc_info.idc_admin_email}
            data = json.dumps(info_json)

        return HttpResponse(data)


    def delete(self,request):
        """删除权限"""
        req_info = eval(request.body.decode())
        idc_id = req_info.get("idc_id")
        asset_db.IDC.objects.get(id=idc_id).delete()
        data = "IDC 已删除,请刷新查看！"
        return HttpResponse(data)


class HostGroup(View):
    """分组管理"""
    @method_decorator(csrf_exempt)
    @method_decorator(login_check)
    @method_decorator(perms_check)
    def dispatch(self, request, *args, **kwargs):
        return super(HostGroup,self).dispatch(request, *args, **kwargs)

    def get(self,request):
        title = "主机分组"
        group_obj = asset_db.HostGroup.objects.all()
        return render(request,'asset_group.html',locals())


    def post(self,request):
        group_name = request.POST.get("group_name")
        group_msg = request.POST.get("group_msg")
        group_obj = asset_db.HostGroup(host_group_name=group_name,host_group_msg=group_msg)
        group_obj.save()
        data = '分组已添加，请刷新查看！'
        return HttpResponse(data)

    def put(self,request):
        '''修改分组'''
        req_info = eval(request.body.decode())
        group_id = req_info.get("group_id")
        group_name = req_info.get("group_name")
        group_msg = req_info.get("group_msg")
        action = req_info.get("action",None)

        if action:
            """修改group信息"""
            group_obj = asset_db.HostGroup.objects.get(id=group_id)
            group_obj.host_group_name = group_name
            group_obj.host_group_msg = group_msg
            group_obj.save()
            data = "分组已修改，请刷新查看！"
            return HttpResponse(data)
        else:
            """获取修改信息"""
            group_info = asset_db.HostGroup.objects.get(id=group_id)
            info_json = {'group_id': group_info.id, 'group_name': group_info.host_group_name, 'group_msg': group_info.host_group_msg}
            data = json.dumps(info_json)

        return HttpResponse(data)


    def delete(self,request):
        """删除分组"""
        req_info = eval(request.body.decode())
        group_id = req_info.get("group_id")
        asset_db.HostGroup.objects.get(id=group_id).delete()
        data = "分组已删除,请刷新查看！"
        return HttpResponse(data)



class Supplier(View):
    """供应商管理"""
    @method_decorator(csrf_exempt)
    @method_decorator(login_check)
    @method_decorator(perms_check)
    def dispatch(self, request, *args, **kwargs):
        return super(Supplier,self).dispatch(request, *args, **kwargs)

    def get(self,request):
        title = "设备厂商"
        supplier_obj = asset_db.Supplier.objects.all()
        return render(request,'asset_supplier.html',locals())


    def post(self,request):
        supplier = request.POST.get("supplier")
        supplier_head = request.POST.get("supplier_head")
        supplier_head_phone = request.POST.get("supplier_head_phone")
        supplier_head_email = request.POST.get("supplier_head_email")
        supplier_obj = asset_db.Supplier(supplier=supplier,supplier_head=supplier_head,supplier_head_phone=supplier_head_phone,
                               supplier_head_email=supplier_head_email)
        supplier_obj.save()
        data = '供应商已添加，请刷新查看！'
        return HttpResponse(data)

    def put(self,request):
        req_info = eval(request.body.decode())
        supplier_id = req_info.get("supplier_id")
        supplier = req_info.get("supplier")
        supplier_head = req_info.get("supplier_head")
        supplier_head_phone = req_info.get("supplier_head_phone")
        supplier_head_email = req_info.get("supplier_head_email")
        action = req_info.get("action",None)
        if action:
            """修改supplier信息"""
            supplier_obj = asset_db.Supplier.objects.get(id=supplier_id)
            supplier_obj.supplier = supplier
            supplier_obj.supplier_head = supplier_head
            supplier_obj.supplier_head_phone = supplier_head_phone
            supplier_obj.supplier_head_email = supplier_head_email
            supplier_obj.save()
            data = "供应商已修改，请刷新查看！"
            return HttpResponse(data)
        else:
            """获取修改信息"""
            supplier_info = asset_db.Supplier.objects.get(id=supplier_id)

            info_json = {'supplier_id': supplier_info.id, 'supplier': supplier_info.supplier,'supplier_head': supplier_info.supplier_head,
                         'supplier_head_phone': supplier_info.supplier_head_phone,'supplier_head_email': supplier_info.supplier_head_email}
            data = json.dumps(info_json)

        return HttpResponse(data)


    def delete(self,request):
        """删除权限"""
        req_info = eval(request.body.decode())
        supplier_id = req_info.get("supplier_id")
        asset_db.Supplier.objects.get(id=supplier_id).delete()
        data = "supplier 已删除,请刷新查看！"
        return HttpResponse(data)



class Host(View):
    """服务器管理"""
    @method_decorator(csrf_exempt)
    @method_decorator(login_check)
    @method_decorator(perms_check)
    def dispatch(self, request, *args, **kwargs):
        return super(Host,self).dispatch(request, *args, **kwargs)

    def get(self,request):
        title = "服务器"
        supplier_obj = asset_db.Supplier.objects.all()
        group_obj = asset_db.HostGroup.objects.all()
        idc_obj = asset_db.IDC.objects.all()

        role_id = request.session['role_id']

        role_obj = auth_db.Role.objects.get(id=role_id)

        host_obj = role_obj.host.all()
        host_list = []

        for i in host_obj:
            try:
                host_status = asset_db.HostDetail.objects.get(host_id=i.id).host_status
            except:
                host_status = "Unknown"
            host_list.append({"id":i.id,"host_ip":i.host_ip,"host_type":i.host_type,"group_name":i.group.host_group_name,"host_msg":i.host_msg,
                              "supplier":i.supplier.supplier,"idc_name":i.idc.idc_name,"host_status":host_status})

        webssh_url = WEBSSH_URL

        return render(request,'asset_host.html',locals())

    def post(self,request):
        host_ip = request.POST.get("host_ip")
        host_remove_port = request.POST.get("host_remove_port")
        host_user = request.POST.get("host_user")
        host_passwd = request.POST.get("host_passwd")
        host_type = request.POST.get("host_type")
        host_group = request.POST.get("host_group")
        host_idc = request.POST.get("host_idc")
        host_supplier = request.POST.get("host_supplier")
        host_msg = request.POST.get("host_msg")
        serial_num = request.POST.get("serial_num")
        purchase_date = request.POST.get("purchase_date")
        overdue_date = request.POST.get("overdue_date")

        if host_group == "0":
            host_group = None

        if host_idc == "0":
            host_idc = None

        if host_supplier == "0":
            host_supplier = None

        # 加密密码
        key = SECRET_KEY[2:18]
        pc = encryption.prpcrypt(key)  # 初始化密钥
        aes_passwd = pc.encrypt(host_passwd)

        host_obj = asset_db.Host(host_ip=host_ip,host_remove_port=host_remove_port,host_user=host_user,host_passwd=aes_passwd,host_type=host_type,
                                 group_id=host_group,idc_id=host_idc,supplier_id=host_supplier,host_msg=host_msg,serial_num=serial_num,
                                 purchase_date=purchase_date,overdue_date=overdue_date)
        host_obj.save()
        data = '服务器已添加，请刷新查看！'
        return HttpResponse(data)

    def put(self,request):
        req_info = eval(request.body.decode())
        host_id = req_info.get("host_id")
        host_ip = req_info.get("host_ip")
        host_remove_port = req_info.get("host_remove_port")
        host_user = req_info.get("host_user")
        host_passwd = req_info.get("host_passwd")
        host_type = req_info.get("host_type")
        host_group = req_info.get("host_group")
        host_idc = req_info.get("host_idc")
        host_supplier = req_info.get("host_supplier")
        host_msg = req_info.get("host_msg")
        serial_num = req_info.get("serial_num")
        purchase_date = req_info.get("purchase_date")
        overdue_date = req_info.get("overdue_date")
        action = req_info.get("action",None)
        if action:
            """修改服务器信息"""
            # 加密密码
            key = SECRET_KEY[2:18]
            pc = encryption.prpcrypt(key)  # 初始化密钥
            aes_passwd = pc.encrypt(host_passwd)

            if host_group == "0":
                host_group = None

            if host_idc == "0":
                host_idc = None

            if host_supplier == "0":
                host_supplier = None

            host_obj = asset_db.Host.objects.get(id=host_id)
            host_obj.host_ip = host_ip
            host_obj.host_remove_port = host_remove_port
            host_obj.host_user = host_user
            host_obj.host_passwd = aes_passwd
            host_obj.host_type = host_type
            host_obj.group_id = host_group
            host_obj.idc_id = host_idc
            host_obj.supplier_id = host_supplier
            host_obj.host_msg = host_msg
            host_obj.serial_num = serial_num
            host_obj.purchase_date = purchase_date
            host_obj.overdue_date = overdue_date

            host_obj.save()
            data = "服务器已修改，请刷新查看！"
            return HttpResponse(data)
        else:
            """获取修改信息"""
            host_info = asset_db.Host.objects.get(id=host_id)

            #密码解密
            key = SECRET_KEY[2:18]
            pc = encryption.prpcrypt(key)
            passwd =host_info.host_passwd.strip("b").strip("'").encode(encoding="utf-8")
            de_passwd = pc.decrypt(passwd).decode()


            info_json = {'host_id': host_info.id, 'host_ip': host_info.host_ip,'host_remove_port': host_info.host_remove_port,
                         'host_user': host_info.host_user,'host_passwd': de_passwd,'host_type': host_info.host_type,
                         'host_group': host_info.group_id,'host_idc': host_info.idc_id,'host_supplier': host_info.supplier_id,
                         'host_msg': host_info.host_msg,'serial_num': host_info.serial_num,'purchase_date': host_info.purchase_date,'overdue_date': host_info.overdue_date}
            data = json.dumps(info_json,ensure_ascii=False)

        return HttpResponse(data)

    def delete(self,request):
        """删除服务器"""
        req_info = eval(request.body.decode())
        host_id = req_info.get("host_id")
        asset_db.Host.objects.get(id=host_id).delete()
        data = "服务器已删除,请刷新查看！"
        return HttpResponse(data)


@login_check
@perms_check
def host_detail(request,id):
    """查看服务器详细信息"""
    title = "服务器"
    host_obj = asset_db.Host.objects.get(id=id)
    try:
        host_detail = asset_db.HostDetail.objects.get(host_id=id)

        disk_info = json.loads(host_detail.disk_info)
        interface = json.loads(host_detail.interface)

        software_obj = asset_db.software.objects.filter(host_id=id)

        software_list = []
        for i in software_obj:
            software_list.append(
                {"server_name": i.server_name, "server_version": i.server_version, "server_port": i.server_port})


        return render(request, "asset_host_detail.html", locals())
    except:
        return HttpResponse("信息未同步")


@csrf_exempt
@login_check
@perms_check
def sync_host_info(request):
    """同步服务器系统信息"""
    ids = request.POST.get("ids")
    ips = []

    if ids == 'all':
        host_obj =  asset_db.Host.objects.all()
        for i in host_obj:
            ips.append(i.host_ip)
    else:
        ids = ids.strip(',').split(',')
        for i in ids:
            host_obj = asset_db.Host.objects.get(id=i)
            ips.append(host_obj.host_ip)


    tk = sync_host.delay(json.dumps(ips),json.dumps(SALT_API),json.dumps(SERVER_TAG))


    data = "信息同步中,任务ID:{}".format(tk.id)

    user_name = request.session['user_name']

    user_obj = auth_db.User.objects.get(user_name=user_name)

    task_obj = log_db.TaskRecord(task_name="同步服务器信息",task_id=tk.id,status=tk.state,task_user_id=user_obj.id)

    task_obj.save()

    return  HttpResponse(data)


@csrf_exempt
@login_check
@perms_check
def search_host(request):
    """过滤服务器信息"""
    idc_id = request.POST.get('idc_id',None)
    group_id = request.POST.get('hostgroup_id',None)
    host_type = request.POST.get('host_type',None)
    search_key = request.POST.get('search_key', None)
    role_id = request.session['role_id']

    if search_key:
        host_obj = asset_db.Host.objects.filter((Q(host_ip__icontains=search_key) | Q(host_msg__icontains=search_key) | Q(host_type__icontains=search_key))& Q(role__id=role_id))

    if idc_id:

        host_obj = asset_db.Host.objects.filter(Q(idc_id=idc_id) & Q(role__id=role_id))

    if group_id:
        host_obj = asset_db.Host.objects.filter(Q(group_id=group_id) & Q(role__id=role_id))

    if host_type:
        host_obj = asset_db.Host.objects.filter(Q(host_type=host_type) & Q(role__id=role_id))

    host_list = []

    for i in host_obj:
        try:
            host_status = asset_db.HostDetail.objects.get(host_id=i.id).host_status
        except:
            host_status = "Unknown"

        host_list.append({"id": i.id, "host_ip": i.host_ip, "host_type": i.host_type, "group_name": i.group.host_group_name,
             "host_msg": i.host_msg,
             "supplier": i.supplier.supplier, "idc_name": i.idc.idc_name, "host_status": host_status})


    return render(request, "asset_host_search.html", locals())



@csrf_exempt
@login_check
@perms_check
def del_host(request):
    """批量删除服务器"""
    ids = request.POST.get("ids")
    ids = ids.strip(',').split(',')
    for ids in ids:
        asset_db.Host.objects.get(id=ids).delete()

    return HttpResponse("服务器已删除,请刷新页面")



@csrf_exempt
@login_check
@perms_check
def connect_host(request):
    """连接服务器"""

    req_info = eval(request.body.decode())

    host_id = req_info.get("host_id")

    host_obj = asset_db.Host.objects.get(id=host_id)

    redis_obj = redis.Redis(host=REDIS_INFO["host"],port=REDIS_INFO["port"])

    remote_user = request.session['remote_user']
    remote_passwd = request.session['remote_passwd']
    remote_sshkey = request.session['remote_sshkey']
    remote_sshkey_pass = request.session['remote_sshkey_pass']
    if remote_sshkey:
        if remote_sshkey_pass:
            pass
        else:
            remote_sshkey_pass = "None"
    else:
        remote_sshkey = "None"
        remote_sshkey_pass = "None"

    webssh_info = {"username": remote_user, "publickey": remote_sshkey, "password": remote_passwd,"hostname": host_obj.host_ip,
                   "port": host_obj.host_remove_port,"key_pass":remote_sshkey_pass}

    redis_obj.set("webssh_info",json.dumps(webssh_info),ex=5,nx=True)

    return HttpResponse("已连接到服务器")


@csrf_exempt
@login_check
@perms_check
def import_host(request):
    """导入服务器信息"""
    upload_file = request.FILES.get("upload_file", None)

    filename = os.path.join(BASE_DIR,"statics/media/import_asset.xlsx")

    file_obj = open(filename,'wb')

    for chrunk in upload_file.chunks():
        file_obj.write(chrunk)
        file_obj.close()

    data = read_excel.import_host(filename)

    for i in data:
        host_ip = i[0]
        host_idc = i[1]
        host_type = i[2]
        host_group = i[3]
        host_user = i[4]
        host_passwd = i[5]
        host_msg = i[6]
        host_remove_port = i[7]
        serial_num = i[8]
        purchase_date = i[9]
        overdue_date = i[10]
        host_supplier = i[11]

        supplier_head = i[12]
        supplier_head_phone = i[13]
        supplier_head_email = i[14]

        try:
            idc_obj = asset_db.IDC.objects.get(idc_name=host_idc)

        except:
            idc_obj = asset_db.IDC(idc_name=host_idc)
            idc_obj.save()

        idc_id = idc_obj.id

        try:
            group_obj = asset_db.HostGroup.objects.get(host_group_name=host_group)
        except:
            group_obj = asset_db.HostGroup(host_group_name=host_group)
            group_obj.save()

        group_id = group_obj.id

        try:
            supplier_obj = asset_db.Supplier.objects.get(supplier=host_supplier)

        except:
            supplier_obj = asset_db.Supplier(supplier=host_supplier, supplier_head=supplier_head,
                                                                   supplier_head_phone=supplier_head_phone,
                                                                   supplier_head_email=supplier_head_email)
            supplier_obj.save()

        supplier_id = supplier_obj.id

        # 加密密码
        key = SECRET_KEY[2:18]
        # 初始化密钥
        pc = encryption.prpcrypt(key)
        aes_passwd = pc.encrypt(host_passwd)

        try:
            host_obj = asset_db.Host(host_ip=host_ip, host_remove_port=host_remove_port, host_user=host_user,
                                     host_passwd=aes_passwd, host_type=host_type,
                                     group_id=group_id, idc_id=idc_id, supplier_id=supplier_id, host_msg=host_msg,
                                     serial_num=serial_num,
                                     purchase_date=purchase_date, overdue_date=overdue_date)
            host_obj.save()
        except Exception as e:
            print(e)
            continue

    return redirect('/asset/host/')

@csrf_exempt
@login_check
@perms_check
def export_host(request):
    """批量导出服务器"""
    role_id = request.session['role_id']

    role_obj = auth_db.Role.objects.get(id=role_id)
    role_obj1 = auth_db.User.objects.get(id=role_id)
   
    host_obj = role_obj.host.all()
    host_obj = host_obj.filter()
    data_list = []

    key = SECRET_KEY[2:18]
    pc = encryption.prpcrypt(key)
 
    for i in host_obj:

        if i.group:
            group_name = i.group.host_group_name
        else:
            group_name = ""
        if i.idc:
            idc_name = i.idc.idc_name
        else:
            idc_name = ""
        if i.supplier:
            supplier_head_email = i.supplier.supplier_head_email
            supplier_head_phone = i.supplier.supplier_head_phone
            supplier_head = i.supplier.supplier_head
            supplier = i.supplier.supplier
        else:
            supplier_head_email = ""
            supplier_head_phone = ""
            supplier_head = ""
            supplier = ""

        passwd = pc.decrypt(i.host_passwd.strip("b").strip("'").encode(encoding="utf-8")).decode()

        host_info = [i.host_ip,idc_name,i.host_type,group_name,i.host_user,
                     passwd,i.host_msg,i.host_remove_port,i.serial_num,i.purchase_date,i.overdue_date,supplier,supplier_head,
                     supplier_head_phone,supplier_head_email,role_obj1.ready_name,role_obj.role_msg,role_obj1.phone,role_obj1.email]
        
        data_list.append(host_info)
  
       

    def headStyle():
        ft = Font(size=14,name='SimSun')  # color="0F0F0F"字体颜色,italic=False,是否斜体,字体样式,大小,bold=False是否粗体

        align = Alignment(horizontal="center", vertical="center")  # 居中对齐

        fill = PatternFill("solid", fgColor="FFFF00")  # 背景填充颜色

        # fill = GradientFill(stop=("000000", "FFFFFF")) #背景填充颜色，渐变

        bd = Side(style='thin', color="000000")

        border = Border(left=bd, top=bd, right=bd, bottom=bd)

        mystyle = NamedStyle(name="mystyle")

        mystyle.font = ft
        mystyle.alignment = align
        mystyle.fill = fill
        mystyle.border = border

        return mystyle


    def boderStyle():
        ft = Font(size=12,name='SimSun')  # color="0F0F0F"字体颜色,italic=False,是否斜体,字体样式,大小,bold=False是否粗体

        align = Alignment(horizontal="center", vertical="center")  # 居中对齐

        #fill = PatternFill("solid", fgColor="FFFF00")  # 背景填充颜色

        # fill = GradientFill(stop=("000000", "FFFFFF")) #背景填充颜色，渐变

        bd = Side(style='thin', color="000000")

        border = Border(left=bd, top=bd, right=bd, bottom=bd)

        bdstyle = NamedStyle(name="bdstyle")

        bdstyle.font = ft
        bdstyle.alignment = align
       # mystyle.fill = fill
       # mystyle.border = border

        return bdstyle



    def makeExcel(data_list):
        global statuss
        try:
            base_dir = BASE_DIR

            path_dir = base_dir + '/statics/media/'
       
           
            wb = Workbook()
         
            
            filename = path_dir + 'cmdb.xlsx'

            ws = wb.active

            ws.title = u'资产清单'

            N= 1
            head = ['IP地址','机房','设备类型','主机组','管理用户','用户密码','描述','远程端口','序列号','购买日期','过保日期','厂商','厂商负责人','厂商电话','厂商邮箱','管理员','角色','电话','邮箱']


            data_list.insert(0,head)

            head_Style = headStyle()
            boder_Style = boderStyle()
            for i in data_list:
                n =1

                for j in i:
                    col = get_column_letter(n)
                    ws["%s%d" % (col,N)] = j

                    if N == 1:
                        ws["%s%d" % (col, N)].style = head_Style
                        ws.column_dimensions['%s' % col ].width = 18  # 设置列宽
                    else:
                        ws["%s%d" % (col, N)].style = boder_Style

                    n+=1
                N+=1

            ws.row_dimensions[1].height = 30  #设置行高


            wb.save(filename)
            statuss = 0
        except Exception as e:
            print(e)
            statuss = 1

    makeExcel(data_list)

    if data_list:
        if statuss == 0:
            success = 1
            message = '服务器信息导出成功'
        else:
            success = 0
            message = '服务器信息转表失败'
    else:
        success = 0
        message = '获取服务器信息失败'
    info_json = {'code':success,'message':message}
    return HttpResponse(json.dumps(info_json,ensure_ascii=False), content_type="application/json")

class Netwk(View):
    """网络设备管理"""
    @method_decorator(csrf_exempt)
    @method_decorator(login_check)
    @method_decorator(perms_check)
    def dispatch(self, request, *args, **kwargs):
        return super(Netwk,self).dispatch(request, *args, **kwargs)

    def get(self,request):
        title = "网络设备"
        supplier_obj = asset_db.Supplier.objects.all()
        group_obj = asset_db.HostGroup.objects.all()
        idc_obj = asset_db.IDC.objects.all()

        role_id = request.session['role_id']
        role_obj = auth_db.Role.objects.get(id=role_id)
        netwk_obj = role_obj.netwk.all()

        return render(request,'asset_netwk.html',locals())

    def post(self,request):
        netwk_ip = request.POST.get("netwk_ip")
        netwk_remove_port = request.POST.get("netwk_remove_port")
        netwk_user = request.POST.get("netwk_user")
        netwk_passwd = request.POST.get("netwk_passwd")
        netwk_type = request.POST.get("netwk_type")
        netwk_group = request.POST.get("netwk_group")
        netwk_idc = request.POST.get("netwk_idc")
        netwk_supplier = request.POST.get("netwk_supplier")
        netwk_msg = request.POST.get("netwk_msg")
        serial_num = request.POST.get("serial_num")
        purchase_date = request.POST.get("purchase_date")
        overdue_date = request.POST.get("overdue_date")

        if netwk_group == "0":
            netwk_group = None

        if netwk_idc == "0":
            netwk_idc = None

        if netwk_supplier == "0":
            netwk_supplier = None

        # 加密密码
        key = SECRET_KEY[2:18]
        pc = encryption.prpcrypt(key)  # 初始化密钥
        aes_passwd = pc.encrypt(netwk_passwd)

        netwk_obj = asset_db.Netwk(netwk_ip=netwk_ip,netwk_remove_port=netwk_remove_port,netwk_user=netwk_user,netwk_passwd=aes_passwd,netwk_type=netwk_type,
                                 group_id=netwk_group,idc_id=netwk_idc,supplier_id=netwk_supplier,netwk_msg=netwk_msg,serial_num=serial_num,
                                 purchase_date=purchase_date,overdue_date=overdue_date)
        netwk_obj.save()
        data = '设备已添加，请刷新查看！'
        return HttpResponse(data)

    def put(self,request):
        req_info = eval(request.body.decode())
        netwk_id = req_info.get("netwk_id")
        netwk_ip = req_info.get("netwk_ip")
        netwk_remove_port = req_info.get("netwk_remove_port")
        netwk_user = req_info.get("netwk_user")
        netwk_passwd = req_info.get("netwk_passwd")
        netwk_type = req_info.get("netwk_type")
        netwk_group = req_info.get("netwk_group")
        netwk_idc = req_info.get("netwk_idc")
        netwk_supplier = req_info.get("netwk_supplier")
        netwk_msg = req_info.get("netwk_msg")
        serial_num = req_info.get("serial_num")
        purchase_date = req_info.get("purchase_date")
        overdue_date = req_info.get("overdue_date")
        action = req_info.get("action",None)
        if action:
            """修改网络设备信息"""
            # 加密密码
            key = SECRET_KEY[2:18]
            pc = encryption.prpcrypt(key)  # 初始化密钥
            aes_passwd = pc.encrypt(netwk_passwd)

            if netwk_group == "0":
                netwk_group = None

            if netwk_idc == "0":
                netwk_idc = None

            if netwk_supplier == "0":
                netwk_supplier = None

            netwk_obj = asset_db.Netwk.objects.get(id=netwk_id)
            netwk_obj.netwk_ip = netwk_ip
            netwk_obj.netwk_remove_port = netwk_remove_port
            netwk_obj.netwk_user = netwk_user
            netwk_obj.netwk_passwd = aes_passwd
            netwk_obj.netwk_type = netwk_type
            netwk_obj.group_id = netwk_group
            netwk_obj.idc_id = netwk_idc
            netwk_obj.supplier_id = netwk_supplier
            netwk_obj.netwk_msg = netwk_msg
            netwk_obj.serial_num = serial_num
            netwk_obj.purchase_date = purchase_date
            netwk_obj.overdue_date = overdue_date
            netwk_obj.save()
            data = "设备已修改，请刷新查看！"
            return HttpResponse(data)
        else:
            """获取修改信息"""
            netwk_info = asset_db.Netwk.objects.get(id=netwk_id)

            #密码解密
            key = SECRET_KEY[2:18]
            pc = encryption.prpcrypt(key)
            passwd =netwk_info.netwk_passwd.strip("b").strip("'").encode(encoding="utf-8")
            de_passwd = pc.decrypt(passwd).decode()


            info_json = {'netwk_id': netwk_info.id, 'netwk_ip': netwk_info.netwk_ip,'netwk_remove_port': netwk_info.netwk_remove_port,
                         'netwk_user': netwk_info.netwk_user,'netwk_passwd': de_passwd,'netwk_type': netwk_info.netwk_type,
                         'netwk_group': netwk_info.group_id,'netwk_idc': netwk_info.idc_id,'netwk_supplier': netwk_info.supplier_id,
                         'netwk_msg': netwk_info.netwk_msg,'serial_num': netwk_info.serial_num,'purchase_date': netwk_info.purchase_date,'overdue_date': netwk_info.overdue_date}
            data = json.dumps(info_json,ensure_ascii=False)

        return HttpResponse(data)

    def delete(self,request):
        """删除网络设备"""
        req_info = eval(request.body.decode())
        netwk_id = req_info.get("netwk_id")
        asset_db.Netwk.objects.get(id=netwk_id).delete()
        data = "设备已删除,请刷新查看！"
        return HttpResponse(data)


